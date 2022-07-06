
import openpyxl as xl

wb = xl.load_workbook("file.xlsx")

sheet = wb["Sheet1"]

cell = sheet["a1"]

cell = sheet.cell(1, 1)

print(sheet.max_row)